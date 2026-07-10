import openpyxl
from openpyxl.cell import cell

from PythonBasicsPractise1.loops import i

book= openpyxl.load_workbook("C:\\Amy Documents\\Book1.xlsx")

sheet = book.active
Dict ={}
cell = sheet.cell(row =1, column = 2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

#for rows in range (1,sheet.max_row+1):
 # if sheet.cell(row=i, cloumn=1).value =="Testcase2":
  #   for j in range (1,sheet.max_column+1):
Dict[sheet.cell(row=1, column =j).value]=sheet.cell(row=i,column=j).value